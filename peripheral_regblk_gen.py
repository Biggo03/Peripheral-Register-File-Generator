#!/usr/bin/env python3
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
import yaml
import sys
import os
import shutil

def process_groups(groups):
    group_headers = [cell.value for cell in next(groups.iter_rows(min_row=1, max_row=1))]
    group_data = {}

    for row in groups.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(group_headers, row))
        group_name = entry.pop("GROUP")

        group_data[group_name] = entry

    return group_data

def process_registers(registers):
    reg_data = {}
    reg_info = {}
    active_reg = ""
    active_group = ""
    offset_max = 0

    reg_headers = [cell.value for cell in next(registers.iter_rows(min_row=1, max_row=1))]
    for row in registers.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(reg_headers, row))
        addr_offset = entry.pop("ADDR_OFFSET")

        # All 32-bit registers have an address offset
        # Fields do not
        if (addr_offset):
            # Complete entry for previous register
            if (active_reg):
                reg_data[active_group][active_reg] = reg_info

            # Extract relavent registers
            offset_max = int(addr_offset, 16)
            active_reg = entry.pop("NAME").upper()
            active_group = entry.pop("ACCESS/GROUP")
            reg_description = entry.pop("REG_DESCRIPTION")

            # Initialize for next register
            reg_data.setdefault(active_group, {})
            reg_data[active_group][active_reg] = {}
            reg_info = {}

            # Store relavent info
            reg_info["FIELDS"] = {}
            reg_info["REG_DESCRIPTION"] = reg_description
            reg_info["ADDR_OFFSET"] = addr_offset
            reg_info["ADDR_MACRO"] = f"{active_reg.upper()}_ADDR"
        else:
            field = entry.pop("NAME").upper()
            # Don't generate anything for reserved fields
            if (field == "RESERVED"):
                continue

            # Extract relavent fields
            field_info = {}
            access = entry.pop("ACCESS/GROUP").upper()
            bits = entry.pop("BITS")
            reset_val = entry.pop("RESET_VAL")
            field_description = entry.pop("FIELD_DESCRIPTION")

            # param calculation
            if (":" in bits):
                high_bit, low_bit = bits.replace("[", "").replace("]", "").split(":")
                field_info["WIDTH"] = int(high_bit) - int(low_bit) + 1
            else:
                field_info["WIDTH"] = 1

            field_info["ACCESS"] = access
            field_info["BITS"] = bits
            field_info["RESET_VAL"] = reset_val
            field_info["FIELD_DESCRIPTION"] = field_description

            reg_info["FIELDS"][field] = field_info

    # Write info for last register
    reg_data[active_group][active_reg] = reg_info
    reg_data["ADDR_WIDTH"] = int(offset_max).bit_length()

    return reg_data

def excel_to_yaml(input_file, peripheral_name, output_file):
    wb = load_workbook(input_file)
    registers = wb["Registers"]
    groups = wb["Groups"]

    group_data = process_groups(groups)
    reg_data = process_registers(registers)

    combined_data = {}

    combined_data["GROUPS"] = group_data
    combined_data["REGISTERS"] = reg_data

    #Prefix the macros
    for reg_group, registers in reg_data.items():
        if ("ADDR_WIDTH" not in reg_group):
            for reg_name, reg_info in registers.items():
                reg_data[reg_group][reg_name]["ADDR_MACRO"] = f'{peripheral_name.upper()}_{reg_info["ADDR_MACRO"]}'


    with open(output_file, "w") as f:
        yaml_str = yaml.dump(combined_data, sort_keys=False)
        yaml_str = yaml_str.replace("\n- ", "\n\n- ")

        f.write(yaml_str)

    return

def generate_macros(peripheral_name, yaml_file, macro_dir):
    with open(yaml_file, "r") as f:
        combined_data = yaml.safe_load(f)
        reg_data = combined_data.get("REGISTERS")

    file_name = f"{peripheral_name}_reg_macros.sv"
    macro_path = os.path.abspath(f"{macro_dir}/{file_name}")

    with open(macro_path, "w") as f:
        f.write("////////////////////////////////////////////////\n")
        f.write("// AUTO-GENERATED PERIPHERAL REGISTER MACROS //\n")
        f.write("///////////////////////////////////////////////\n")

        # figure out longest register name for alignment
        max_name_len = max(len(reg_name) for reg_name in reg_data.keys())

        for reg_group in reg_data.keys():
            if (reg_group == "ADDR_WIDTH"):
                continue
            group_header = f"// REGISTER GROUP: {reg_group}\n"
            f.write(group_header)
            for reg_info in reg_data[reg_group].values():
                name_field = reg_info['ADDR_MACRO'].ljust(max_name_len + 10)
                addr_field = f"{reg_data['ADDR_WIDTH']}'h{reg_info['ADDR_OFFSET'].replace('0x', '')}"
                line = f"`define {name_field} {addr_field} // {reg_info['REG_DESCRIPTION']}\n"
                f.write(line)
            f.write("\n")

    shutil.copy(macro_path, f"./outputs/{file_name}")

    return macro_path

def generate_package(peripheral_name, yaml_file, package_op_dir):
    with open(yaml_file, "r") as f:
        combined_data = yaml.safe_load(f)
        reg_data = combined_data.get("REGISTERS")

    file_name = f"{peripheral_name}_reg_package.sv"
    package_path = os.path.abspath(f"{package_op_dir}/{file_name}")

    # Jinja setup
    env = Environment(loader=FileSystemLoader("templates"))

    package_template = env.get_template("reg_package_template.sv.j2")
    package_output = package_template.render(peripheral_name=peripheral_name,
                                             reg_data=reg_data)

    with open(package_path, "w") as f:
        f.write(package_output)

    shutil.copy(package_path, f"./outputs/{file_name}")

    return package_path

def generate_rtl(peripheral_name, yaml_file, rtl_op_dir, tb_op_dir):
    with open(yaml_file, "r") as f:
        combined_data = yaml.safe_load(f)
        reg_data = combined_data.get("REGISTERS")
        group_data = combined_data.get("GROUPS")

    rtl_file_name = f"{peripheral_name}_regfile.sv"
    rtl_path = os.path.abspath(f"{rtl_op_dir}/{rtl_file_name}")

    tb_file_name = f"{peripheral_name}_regfile_tb.sv"
    tb_path = os.path.abspath(f"{tb_op_dir}/{tb_file_name}")

    for reg_group in reg_data.keys():
        if (reg_group == "ADDR_WIDTH"):
            continue
        for reg_name, reg_info in reg_data[reg_group].items():
            reg_data[reg_group][reg_name]["ADDR_MACRO"] = f"`{reg_info['ADDR_MACRO']}"
            for field_name, field_info in reg_info["FIELDS"].items():
                reg_data[reg_group][reg_name]["FIELDS"][field_name]["RESET_VAL"] = f"{field_info['WIDTH']}'h{field_info['RESET_VAL'].replace('x0', '')}"


    max_name_len = max(len(name) for name in reg_data.keys())

    # Jinja setup
    env = Environment(loader=FileSystemLoader("templates"))

    rtl_template = env.get_template("regfile_template.sv.j2")
    rtl_output = rtl_template.render(peripheral_name=peripheral_name,
                                     reg_data=reg_data,
                                     group_data=group_data,
                                     max_name_len=max_name_len)

    tb_template = env.get_template("regfile_tb_template.sv.j2")
    tb_output = tb_template.render(peripheral_name=peripheral_name,
                                   reg_data=reg_data,
                                   group_data=group_data,
                                   max_name_len=max_name_len)

    with open(rtl_path, "w") as rtl_file, open(tb_path, "w") as tb_file:
        rtl_file.write(rtl_output)
        tb_file.write(tb_output)

    shutil.copy(rtl_path, f"./outputs/{rtl_file_name}")
    shutil.copy(tb_path, f"./outputs/{tb_file_name}")

    return rtl_path,tb_path

def generate_c_defs(peripheral_name, yaml_file, cdef_op_dir):
    with open(yaml_file, "r") as f:
        combined_data = yaml.safe_load(f)
        reg_data = combined_data.get("REGISTERS")

    c_macro_file_name = f"{peripheral_name}_reg_macros.h"

    define_info = {}
    max_name_len = 0
    for group_name, group_info in reg_data.items():
        if (group_name == "ADDR_WIDTH"):
            continue
        for reg_name, reg_info in group_info.items():
            if (len(reg_info["ADDR_MACRO"]) > max_name_len):
                max_name_len = len(reg_info["ADDR_MACRO"])


            addr_offset = int(reg_info["ADDR_OFFSET"][2:]) & 0xFFFFFFFF
            define_info[f"{reg_name.upper()}"] = {
                "ADDR_MACRO": reg_info["ADDR_MACRO"],
                "ADDR_OFFSET": f"0x{addr_offset:08X}u",
                "REG_DESCRIPTION": reg_info["REG_DESCRIPTION"]
            }

            for field_name, field_info in reg_info["FIELDS"].items():
                field_dict = {}
                field_dict["MASK"] = range_to_mask(field_info["BITS"])
                field_dict["FIELD_DESCRIPTION"] = field_info["FIELD_DESCRIPTION"]

                define_info[reg_name.upper()][field_name] = field_dict

                if (len(field_name) > max_name_len):
                    max_name_len = len(field_name)

    env = Environment(loader=FileSystemLoader("templates"))

    c_macro_template = env.get_template("reg_c_macros.h.j2")
    reg_c_macros_output = c_macro_template.render(peripheral_name=peripheral_name,
                                                  define_info=define_info,
                                                  max_name_len=max_name_len)

    with open(f"{cdef_op_dir}/{c_macro_file_name}", "w") as cdef_file:
        cdef_file.write(reg_c_macros_output)

    shutil.copy(f"{cdef_op_dir}/{c_macro_file_name}", f"./outputs/{c_macro_file_name}")

def range_to_mask(bit_range):
    # Expect "[msb:lsb]"
    # Or [bit]
    if (":" in bit_range):
        msb_str, lsb_str = bit_range.strip("[]").split(":")
        msb = int(msb_str)
        lsb = int(lsb_str)
    else:
        msb = int(bit_range.strip("[]"))
        lsb = int(bit_range.strip("[]"))

    if msb < lsb:
        raise ValueError(f"Invalid bit range {bit_range}")

    width = msb - lsb + 1
    mask = ((1 << width) - 1) << lsb

    mask_hex = f"0x{(mask & 0xFFFFFFFF):08X}u"

    return mask_hex

def generate_run_script(peripheral_name, script_output_dir, sim_output_dir, macro_path, package_path, rtl_path, tb_path):
    script_path = f"{script_output_dir}/run_regfile_tb.sh"
    sim_output_dir = os.path.abspath(sim_output_dir)

    # Jinja setup
    env = Environment(loader=FileSystemLoader("templates"))

    run_script_template = env.get_template("run_regfile_tb_template.sh.j2")
    run_script_output = run_script_template.render(peripheral_name=peripheral_name,
                                                   sim_output_dir=sim_output_dir,
                                                   macro_path=macro_path,
                                                   package_path=package_path,
                                                   tb_path=tb_path,
                                                   rtl_path=rtl_path)

    with open(script_path, "w") as f:
        f.write(run_script_output)

def main():
    if len(sys.argv) != 2:
        print("Usage: python3 peripheral_regblk_gen.py <yaml_config_file>")
        sys.exit(1)

    config_file = sys.argv[1]

    with open(config_file, "r") as f:
        path_info = yaml.safe_load(f)

    peripheral_name = path_info.get("peripheral_name", "no_name_provided")
    spreadsheet_path = path_info.get("spreadsheet_path", "./outputs")
    rtl_output_dir = path_info.get("rtl_output_dir", "./outputs")
    tb_output_dir = path_info.get("tb_output_dir", "./outputs")
    package_output_dir = path_info.get("package_output_dir", "./outputs")
    macro_output_dir = path_info.get("macro_output_dir", "./outputs")
    cdef_output_dir = path_info.get("cdef_output_dir", "./outputs")
    script_output_dir = path_info.get("script_output_dir", "./outputs")
    sim_output_dir = path_info.get("sim_output_dir", "./outputs")

    yaml_file = f"./outputs/{peripheral_name}_registers.yml"

    os.makedirs(rtl_output_dir, exist_ok=True)
    os.makedirs(tb_output_dir, exist_ok=True)
    os.makedirs(package_output_dir, exist_ok=True)
    os.makedirs(package_output_dir, exist_ok=True)
    os.makedirs(macro_output_dir, exist_ok=True)
    os.makedirs(cdef_output_dir, exist_ok=True)
    os.makedirs("./outputs", exist_ok=True)

    excel_to_yaml(spreadsheet_path, peripheral_name, yaml_file)
    macro_path = generate_macros(peripheral_name, yaml_file, macro_output_dir)
    package_path = generate_package(peripheral_name, yaml_file, package_output_dir)
    rtl_path, tb_path = generate_rtl(peripheral_name, yaml_file, rtl_output_dir, tb_output_dir)
    generate_run_script(peripheral_name, script_output_dir, sim_output_dir, macro_path, package_path, rtl_path, tb_path)

    generate_c_defs(peripheral_name, yaml_file, cdef_output_dir)

if __name__ == "__main__":
    main()
